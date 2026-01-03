#!/usr/bin/env python3
"""
Excel公式重新计算脚本
使用LibreOffice重新计算Excel文件中的所有公式
"""

import json
import sys
import subprocess
import os
import platform
from pathlib import Path
from openpyxl import load_workbook


def setup_libreoffice_macro():
    """如果尚未配置，则设置用于重新计算的LibreOffice宏"""
    if platform.system() == 'Darwin':
        macro_dir = os.path.expanduser('~/Library/Application Support/LibreOffice/4/user/basic/Standard')
    else:
        macro_dir = os.path.expanduser('~/.config/libreoffice/4/user/basic/Standard')
    
    macro_file = os.path.join(macro_dir, 'Module1.xba')
    
    if os.path.exists(macro_file):
        with open(macro_file, 'r') as f:
            if 'RecalculateAndSave' in f.read():
                return True
    
    if not os.path.exists(macro_dir):
        subprocess.run(['soffice', '--headless', '--terminate_after_init'], 
                      capture_output=True, timeout=10)
        os.makedirs(macro_dir, exist_ok=True)
    
    macro_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''
    
    try:
        with open(macro_file, 'w') as f:
            f.write(macro_content)
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    """
    重新计算Excel文件中的公式并报告任何错误
    
    参数:
        filename: Excel文件的路径
        timeout: 等待重新计算的最大时间（秒）
    
    返回:
        包含错误位置和计数的字典
    """
    if not Path(filename).exists():
        return {'error': f'文件 {filename} 不存在'}
    
    abs_path = str(Path(filename).absolute())
    
    if not setup_libreoffice_macro():
        return {'error': '设置LibreOffice宏失败'}
    
    cmd = [
        'soffice', '--headless', '--norestore',
        'vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application',
        abs_path
    ]
    
    # 处理Linux和macOS之间的超时命令差异
    if platform.system() != 'Windows':
        timeout_cmd = 'timeout' if platform.system() == 'Linux' else None
        if platform.system() == 'Darwin':
            # 检查macOS上是否有gtimeout可用
            try:
                subprocess.run(['gtimeout', '--version'], capture_output=True, timeout=1, check=False)
                timeout_cmd = 'gtimeout'
            except (FileNotFoundError, subprocess.TimeoutExpired):
                pass
        
        if timeout_cmd:
            cmd = [timeout_cmd, str(timeout)] + cmd
    
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    if result.returncode != 0 and result.returncode != 124:  # 124是超时退出代码
        error_msg = result.stderr or '重新计算期间发生未知错误'
        if 'Module1' in error_msg or 'RecalculateAndSave' not in error_msg:
            return {'error': 'LibreOffice宏配置不正确'}
        else:
            return {'error': error_msg}
    
    # 检查重新计算文件中的Excel错误 - 扫描所有单元格
    try:
        wb = load_workbook(filename, data_only=True)
        
        excel_errors = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
        error_details = {err: [] for err in excel_errors}
        total_errors = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 检查所有行和列 - 无限制
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break
        
        wb.close()
        
        # 构建结果摘要
        result = {
            'status': 'success' if total_errors == 0 else 'errors_found',
            'total_errors': total_errors,
            'error_summary': {}
        }
        
        # 添加非空错误类别
        for err_type, locations in error_details.items():
            if locations:
                result['error_summary'][err_type] = {
                    'count': len(locations),
                    'locations': locations[:20]  # 显示最多20个位置
                }
        
        # 添加公式计数作为上下文 - 同样检查所有单元格
        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
        wb_formulas.close()
        
        result['total_formulas'] = formula_count
        
        return result
        
    except Exception as e:
        return {'error': str(e)}


def main():
    if len(sys.argv) < 2:
        print("用法: python recalc.py <excel_file> [timeout_seconds]")
        print("\n使用LibreOffice重新计算Excel文件中的所有公式")
        print("\n返回包含错误详情的JSON:")
        print("  - status: 'success' 或 'errors_found'")
        print("  - total_errors: 发现的Excel错误总数")
        print("  - total_formulas: 文件中的公式数量")
        print("  - error_summary: 按错误类型分类的详细信息及位置")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)
    
    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    
    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == '__main__':
    main()